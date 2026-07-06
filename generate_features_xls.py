import pandas as pd

rows = [
    # ── VARIABLE CIBLE ──────────────────────────────────────────────────────────
    ("Cible", "label_failure_next_6h",  "Binaire", "0 / 1",
     "Indique si un incident de sévérité ≥ 4 survient dans les 6 prochaines heures. "
     "Construit en regardant en avant dans la chronologie : 1 si au moins un incident critique "
     "est enregistré dans la fenêtre [t+1h, t+6h], 0 sinon."),

    ("Cible", "label_failure_next_12h", "Binaire", "0 / 1",
     "Même logique que label_failure_next_6h mais sur une fenêtre de 12 heures."),

    ("Cible", "label_failure_next_24h", "Binaire", "0 / 1",
     "Même logique sur 24 heures. Horizon le plus souvent utilisé comme cible principale "
     "dans les modèles entraînés."),

    ("Cible", "label_failure_next_48h", "Binaire", "0 / 1",
     "Même logique sur 48 heures. Horizon le plus large — plus facile à prédire mais moins "
     "actionnable opérationnellement."),

    # ── AGRÉGATS 1H ─────────────────────────────────────────────────────────────
    ("Agrégat 1h", "temperature_c",         "Numérique", "°C",
     "Température moyenne de la machine sur l'heure courante (agrégat des mesures brutes capteur)."),
    ("Agrégat 1h", "temperature_c_mean",    "Numérique", "°C",
     "Synonyme de temperature_c — moyenne horaire explicitement nommée."),
    ("Agrégat 1h", "temperature_c_max",     "Numérique", "°C",
     "Valeur maximale de température enregistrée sur l'heure. Utile pour détecter les pics thermiques brefs."),
    ("Agrégat 1h", "temperature_c_std",     "Numérique", "°C",
     "Écart-type de la température sur l'heure. Une forte dispersion indique une instabilité thermique."),

    ("Agrégat 1h", "pressure_bar",          "Numérique", "bar",
     "Pression moyenne sur l'heure courante."),
    ("Agrégat 1h", "pressure_bar_mean",     "Numérique", "bar",
     "Moyenne horaire explicitement nommée de la pression."),
    ("Agrégat 1h", "pressure_bar_max",      "Numérique", "bar",
     "Pression maximale sur l'heure — détecte les surpressions transitoires."),
    ("Agrégat 1h", "pressure_bar_std",      "Numérique", "bar",
     "Écart-type de la pression sur l'heure. Variabilité élevée = comportement anormal."),

    ("Agrégat 1h", "voltage_mean_v",        "Numérique", "V",
     "Tension électrique moyenne sur l'heure."),
    ("Agrégat 1h", "voltage_mean_v_mean",   "Numérique", "V",
     "Redondance explicite de voltage_mean_v."),
    ("Agrégat 1h", "voltage_mean_v_max",    "Numérique", "V",
     "Tension maximale sur l'heure — pics de tension pouvant signaler des surcharges."),
    ("Agrégat 1h", "voltage_mean_v_std",    "Numérique", "V",
     "Écart-type de la tension — instabilité électrique."),

    ("Agrégat 1h", "rotation_mean_rpm",     "Numérique", "RPM",
     "Vitesse de rotation moyenne sur l'heure."),
    ("Agrégat 1h", "rotation_mean_rpm_mean","Numérique", "RPM",
     "Moyenne horaire de la rotation, nommage explicite."),
    ("Agrégat 1h", "rotation_mean_rpm_max", "Numérique", "RPM",
     "Rotation maximale sur l'heure — sur-régime potentiel."),
    ("Agrégat 1h", "rotation_mean_rpm_std", "Numérique", "RPM",
     "Écart-type de la rotation — irrégularités mécaniques."),

    ("Agrégat 1h", "pieces_produced",       "Numérique", "unités",
     "Nombre de pièces produites sur l'heure — proxy de la charge de travail."),
    ("Agrégat 1h", "pieces_produced_mean",  "Numérique", "unités",
     "Moyenne horaire du nombre de pièces produites."),
    ("Agrégat 1h", "pieces_produced_max",   "Numérique", "unités",
     "Maximum horaire des pièces produites."),
    ("Agrégat 1h", "pieces_produced_std",   "Numérique", "unités",
     "Écart-type — cadence irrégulière pouvant indiquer des micro-arrêts."),

    # ── FENÊTRES GLISSANTES ──────────────────────────────────────────────────────
    ("Fenêtre glissante", "temp_mean_6h",   "Numérique", "°C",
     "Moyenne de la température sur les 6 dernières heures. Lisse les pics ponctuels pour révéler une tendance."),
    ("Fenêtre glissante", "temp_max_6h",    "Numérique", "°C",
     "Température maximale sur 6h — niveau d'exposition au risque thermique récent."),
    ("Fenêtre glissante", "temp_std_6h",    "Numérique", "°C",
     "Variabilité thermique sur 6h — instabilité soutenue."),
    ("Fenêtre glissante", "temp_mean_12h",  "Numérique", "°C",
     "Moyenne de la température sur 12h."),
    ("Fenêtre glissante", "temp_max_12h",   "Numérique", "°C",
     "Maximum sur 12h."),
    ("Fenêtre glissante", "temp_std_12h",   "Numérique", "°C",
     "Écart-type sur 12h."),
    ("Fenêtre glissante", "temp_mean_24h",  "Numérique", "°C",
     "Moyenne sur 24h — baseline journalière de la température."),
    ("Fenêtre glissante", "temp_max_24h",   "Numérique", "°C",
     "Maximum sur 24h."),
    ("Fenêtre glissante", "temp_std_24h",   "Numérique", "°C",
     "Écart-type sur 24h."),

    ("Fenêtre glissante", "pressure_mean_6h",  "Numérique", "bar", "Pression moyenne sur 6h."),
    ("Fenêtre glissante", "pressure_max_6h",   "Numérique", "bar", "Pression maximale sur 6h."),
    ("Fenêtre glissante", "pressure_std_6h",   "Numérique", "bar", "Variabilité de la pression sur 6h."),
    ("Fenêtre glissante", "pressure_mean_12h", "Numérique", "bar", "Pression moyenne sur 12h."),
    ("Fenêtre glissante", "pressure_max_12h",  "Numérique", "bar", "Pression maximale sur 12h."),
    ("Fenêtre glissante", "pressure_std_12h",  "Numérique", "bar", "Variabilité sur 12h."),
    ("Fenêtre glissante", "pressure_mean_24h", "Numérique", "bar", "Pression moyenne sur 24h — baseline journalière."),
    ("Fenêtre glissante", "pressure_max_24h",  "Numérique", "bar", "Maximum sur 24h."),
    ("Fenêtre glissante", "pressure_std_24h",  "Numérique", "bar", "Écart-type sur 24h."),

    ("Fenêtre glissante", "voltage_mean_6h",  "Numérique", "V", "Tension moyenne sur 6h."),
    ("Fenêtre glissante", "voltage_std_6h",   "Numérique", "V", "Variabilité de la tension sur 6h."),
    ("Fenêtre glissante", "voltage_mean_12h", "Numérique", "V", "Tension moyenne sur 12h."),
    ("Fenêtre glissante", "voltage_std_12h",  "Numérique", "V", "Variabilité sur 12h."),
    ("Fenêtre glissante", "voltage_mean_24h", "Numérique", "V", "Tension moyenne sur 24h — baseline journalière."),
    ("Fenêtre glissante", "voltage_std_24h",  "Numérique", "V", "Écart-type sur 24h."),

    ("Fenêtre glissante", "rotation_mean_6h",  "Numérique", "RPM", "Rotation moyenne sur 6h."),
    ("Fenêtre glissante", "rotation_std_6h",   "Numérique", "RPM", "Variabilité de la rotation sur 6h."),
    ("Fenêtre glissante", "rotation_mean_12h", "Numérique", "RPM", "Rotation moyenne sur 12h."),
    ("Fenêtre glissante", "rotation_std_12h",  "Numérique", "RPM", "Variabilité sur 12h."),
    ("Fenêtre glissante", "rotation_mean_24h", "Numérique", "RPM", "Rotation moyenne sur 24h."),
    ("Fenêtre glissante", "rotation_std_24h",  "Numérique", "RPM", "Écart-type sur 24h."),

    ("Fenêtre glissante", "pieces_produced_sum_24h", "Numérique", "unités",
     "Total de pièces produites sur 24h — charge journalière cumulée de la machine."),

    # ── TENDANCES / DELTAS ───────────────────────────────────────────────────────
    ("Tendance / Delta", "temp_delta_1h",      "Numérique", "°C/h",
     "Différence de température entre l'heure courante et il y a 1h. Positif = montée en température."),
    ("Tendance / Delta", "temp_delta_3h",      "Numérique", "°C",
     "Différence sur 3h — tendance à moyen terme."),
    ("Tendance / Delta", "temp_trend_6h",      "Numérique", "°C",
     "Pente estimée (régression linéaire ou différence) de la température sur 6h."),

    ("Tendance / Delta", "pressure_delta_1h",  "Numérique", "bar/h", "Variation de pression sur 1h."),
    ("Tendance / Delta", "pressure_delta_3h",  "Numérique", "bar",   "Variation de pression sur 3h."),
    ("Tendance / Delta", "pressure_trend_6h",  "Numérique", "bar",   "Tendance de la pression sur 6h."),

    ("Tendance / Delta", "voltage_delta_1h",   "Numérique", "V/h",  "Variation de tension sur 1h."),
    ("Tendance / Delta", "voltage_delta_3h",   "Numérique", "V",    "Variation de tension sur 3h."),
    ("Tendance / Delta", "voltage_trend_6h",   "Numérique", "V",    "Tendance de la tension sur 6h."),

    ("Tendance / Delta", "rotation_delta_1h",  "Numérique", "RPM/h", "Variation de rotation sur 1h."),
    ("Tendance / Delta", "rotation_delta_3h",  "Numérique", "RPM",   "Variation de rotation sur 3h."),
    ("Tendance / Delta", "rotation_trend_6h",  "Numérique", "RPM",   "Tendance de la rotation sur 6h."),

    # ── Z-SCORES ─────────────────────────────────────────────────────────────────
    ("Z-score", "temp_zscore_24h",          "Numérique", "σ",
     "Z-score de la température par rapport à la distribution des 24 dernières heures. "
     "Mesure l'anomalie relative à la baseline journalière."),
    ("Z-score", "temp_zscore_machine",      "Numérique", "σ",
     "Z-score de la température par rapport à l'historique complet de la machine. "
     "Capture les déviations structurelles long terme."),
    ("Z-score", "pressure_zscore_24h",      "Numérique", "σ",
     "Z-score de la pression par rapport aux 24 dernières heures."),
    ("Z-score", "pressure_zscore_machine",  "Numérique", "σ",
     "Z-score de la pression par rapport à l'historique machine complet."),

    # ── HISTORIQUE INCIDENTS ─────────────────────────────────────────────────────
    ("Historique incidents", "incident_count_prev_24h",          "Entier", "nb",
     "Nombre d'incidents (toutes sévérités) survenus sur la machine dans les 24h précédentes."),
    ("Historique incidents", "incident_max_severity_prev_24h",   "Numérique", "1–5",
     "Sévérité maximale parmi les incidents des 24 dernières heures. "
     "Proxy de la gravité récente des problèmes."),
    ("Historique incidents", "incident_count_prev_7d",           "Entier", "nb",
     "Nombre total d'incidents sur les 7 derniers jours — fréquence de défaillance hebdomadaire."),
    ("Historique incidents", "hours_since_last_incident",        "Numérique", "h",
     "Nombre d'heures écoulées depuis le dernier incident. "
     "Une valeur faible indique une machine récemment défaillante."),

    ("Type d'incident (24h)", "type_surchauffe_count_prev_24h",          "Entier", "nb",
     "Nombre d'incidents de type 'surchauffe' dans les 24h précédentes."),
    ("Type d'incident (24h)", "type_baisse_pression_count_prev_24h",     "Entier", "nb",
     "Nombre d'incidents 'baisse de pression' dans les 24h."),
    ("Type d'incident (24h)", "type_vibration_count_prev_24h",           "Entier", "nb",
     "Nombre d'incidents 'vibration anormale' dans les 24h."),
    ("Type d'incident (24h)", "type_bruit_mecanique_count_prev_24h",     "Entier", "nb",
     "Nombre d'incidents 'bruit mécanique' dans les 24h."),
    ("Type d'incident (24h)", "type_surconsommation_count_prev_24h",     "Entier", "nb",
     "Nombre d'incidents 'surconsommation électrique' dans les 24h."),
    ("Type d'incident (24h)", "type_blocage_mecanique_count_prev_24h",   "Entier", "nb",
     "Nombre d'incidents 'blocage mécanique' dans les 24h."),
    ("Type d'incident (24h)", "type_alarme_capteur_count_prev_24h",      "Entier", "nb",
     "Nombre d'incidents 'alarme capteur' dans les 24h — peut indiquer un capteur défaillant."),
    ("Type d'incident (24h)", "type_arret_urgence_count_prev_24h",       "Entier", "nb",
     "Nombre d'arrêts d'urgence dans les 24h — événement de sécurité critique."),
    ("Type d'incident (24h)", "type_defaut_qualite_count_prev_24h",      "Entier", "nb",
     "Nombre d'incidents 'défaut qualité' dans les 24h — non-conformités de production."),
]

df = pd.DataFrame(rows, columns=[
    "Catégorie", "Nom de la feature", "Type", "Unité / Valeurs", "Description"
])

out_path = r"c:\Users\Aelion\py-init\features_description.xlsx"

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Features")

    ws = writer.sheets["Features"]
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Couleurs par catégorie ───────────────────────────────────────────────────
    COLORS = {
        "Cible":                  "FF4C4C",   # rouge
        "Agrégat 1h":             "4472C4",   # bleu
        "Fenêtre glissante":      "70AD47",   # vert
        "Tendance / Delta":       "ED7D31",   # orange
        "Z-score":                "9966CC",   # violet
        "Historique incidents":   "00B0F0",   # cyan
        "Type d'incident (24h)": "FFC000",   # jaune doré
    }

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # En-têtes
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Largeurs de colonnes
    col_widths = [22, 38, 12, 18, 90]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30

    # Lignes de données
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cat = row[0].value or ""
        hex_color = COLORS.get(cat, "FFFFFF")
        bg = PatternFill("solid", fgColor=hex_color)
        font_color = "FFFFFF" if cat in ("Cible", "Agrégat 1h", "Historique incidents") else "1F1F1F"

        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column == 1:  # catégorie
                cell.fill = bg
                cell.font = Font(bold=True, color=font_color, size=10)
            elif cell.column == 2:  # nom feature
                cell.font = Font(bold=True, size=10, name="Consolas")
            else:
                cell.font = Font(size=10)

        ws.row_dimensions[row[0].row].height = 40

    ws.freeze_panes = "A2"

print(f"Fichier généré : {out_path}")
